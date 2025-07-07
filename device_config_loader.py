import openpyxl

_selected_device_coords = {}

def load_coordinates_from_default_file():
    path = "coordinates.xlsx"  # Assumes file is in the same directory as the script
    try:
        workbook = openpyxl.load_workbook(path)
    except FileNotFoundError:
        raise FileNotFoundError(f"❌ Could not find '{path}' in the current directory.")

    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    # Collect device names from the Excel
    device_names = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]

    if not device_names:
        raise ValueError("❌ No devices found in Excel.")

    print("\n📱 Available devices:")
    for idx, name in enumerate(device_names, 1):
        print(f"  [{idx}] {name}")

    while True:
        try:
            selected_index = int(input("Enter the number corresponding to your device: "))
            if 1 <= selected_index <= len(device_names):
                device_name = device_names[selected_index - 1]
                break
            else:
                print("⚠️ Invalid selection. Try again.")
        except ValueError:
            print("⚠️ Please enter a number.")

    # Load and return device coordinates
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == device_name:
            global _selected_device_coords
            _selected_device_coords = dict(zip(headers[1:], row[1:]))  # Skip "Device" column
            print(f"\n✅ Coordinates loaded for device: {device_name}")
            return

    raise ValueError(f"❌ Device '{device_name}' not found in Excel.")

def get_device_coords():
    return _selected_device_coords
